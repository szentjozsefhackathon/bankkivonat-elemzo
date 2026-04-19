import xml.etree.ElementTree as ET
import pandas as pd
import argparse, os, ast, time
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from itertools import combinations
from rich.progress import track

# TRANSFORMERS IMPORTS
from transformers import pipeline
from transformers import logging as transformers_logging
from transformers.utils import logging

def get_name_entities(text):
    entities = []
    ner_results= ner_pipeline(text)
    if not ner_results:
        return entities

    # Ha már aggregált a lista (van entity_group kulcs), csak visszaadjuk tisztítva
    if 'entity_group' in ner_results[0]:
        return [{'type': r['entity_group'], 'word': r['word'], 'score': r['score']} for r in ner_results]

    # Ha nem aggregált, manuálisan összefűzzük a start/end indexek alapján
    current_ent = None
    for res in ner_results:
        entity_type = res.get('entity', '').split('-')[-1]
        start, end = res['start'], res['end']
        
        # Ha az előző entitás közvetlenül (vagy 1 karakteren belül) folytatódik
        if entities and entity_type == entities[-1]['type'] and start <= entities[-1]['end'] + 1:
            entities[-1]['end'] = end
        else:
            entities.append({
                'type': entity_type,
                'start': start,
                'end': end
            })

    if current_ent:
        entities.append(current_ent)

    for ent in entities:
        ent['word'] = text[ent['start']:ent['end']].strip()
        
    return entities

def osszeg_kereso(osszeg, darabszam, szamlista):
    for kombi in combinations(szamlista, darabszam):
        if sum(kombi) == osszeg:
            return list(kombi)
    return None

def ujfFeldolgozó(file_path):
    kezdosor = ""
    tranzakciok = []

    with open(file_path, "r") as file:
        elso = True
        for i in file:
            if elso:
                kezdosor = i
                elso = False
            else:
                tranzakciok.append(i)
    
    ossz_osszeg = int(kezdosor[42:54])
    processing_date = f"{kezdosor[-9:-5]}-{kezdosor[-5:-3]}-{kezdosor[-3:]}".strip()    
    visszateres = [processing_date, ossz_osszeg]
    
    tra_list = []
    for i in tranzakciok:
        tra = [f"{i[-16:-12]}-{i[-12:-10]}-{i[-10:-8]}", int(i[61:70])]
        tra_list.append(tra)
    
    visszateres.append(tra_list)
    return visszateres

def megjegyzesKeszito(tomb):
    results = []
    for i in tomb:
        if i["type"] == "PER":
            results.append(i["word"])
    return ", ".join(results)

def get_balance(root, ns):
    opening = 0.0
    closing = 0.0
    for bal in root.findall('.//ns:Stmt/ns:Bal', ns):
        tp_node = bal.find('.//ns:CdOrPrtry/ns:Cd', ns)
        amt = float(bal.find('ns:Amt', ns).text)
        if tp_node is not None:
            if tp_node.text == 'OPBD': # Nyitó könyvelt 
                opening = amt
            elif tp_node.text == 'CLBD': # Záró könyvelt 
                closing = amt
    return opening, closing

def add_new_line(srsz="", date="", biz="", kod="", jogc="", jovir="", ter="", egyenleg="", partner="", kozlemeny="", megj="", ev=""):
    return {
                "Sorszám": srsz,
                "Dátum": date,
                "Bizonylat": biz,
                "Kód": kod,
                "Jogcím": jogc,
                "Jóváírás": jovir,
                "Terhelés": ter,
                "Egyenleg": egyenleg,
                "Partner neve": partner,
                "Közlemény": kozlemeny,
                "Megjegyzés": megj,
                "Egyházadó év": ev
            }

def convert_camt053_to_xlsx(xml_file, output_file, printable, folder_path):
    # XML beolvasása
    tree = ET.parse(xml_file)
    root = tree.getroot()
    ns = {'ns': 'urn:iso:std:iso:20022:tech:xsd:camt.053.001.02'}
    
    # Bizonylatszám/Kivonat szám kinyerése a fejlécből (LglSeqNb (XML tag))
    lgl_seq_node = root.find('.//ns:Stmt/ns:LglSeqNb', ns)
    bizonylat_sorszam = lgl_seq_node.text if lgl_seq_node is not None else ""
    
    # 1. Nyitó és Záró egyenlegek kinyerése a fileból
    opening_balance, closing_balance = get_balance(root, ns)
    current_running_balance = opening_balance
    transactions = []
    sorsz = 1

    # Nyitó Egyenleg Táblázathoz Adása
    transactions.append(add_new_line(egyenleg=opening_balance, megj="Nyitó Egyenleg"))

    # File bejárása Entry-nként
    for ntry in track(root.findall('.//ns:Stmt/ns:Ntry', ns), description="Feldolgozás..."):
        amt_node = ntry.find('ns:Amt', ns)
        amount = float(amt_node.text) if amt_node is not None else 0.0
        indicator = ntry.find('ns:CdtDbtInd', ns).text if ntry.find('ns:CdtDbtInd', ns) is not None else ""
        
        # Jóváírás/Terhelés szétválasztás és egyenleg
        if indicator == "CRDT":
            credit_amount, debit_amount = amount, None
            current_running_balance += amount
        else:
            credit_amount, debit_amount = None, amount
            current_running_balance -= amount
        
        bookg_dt = ntry.find('.//ns:BookgDt/ns:Dt', ns).text if ntry.find('.//ns:BookgDt/ns:Dt', ns) is not None else ""
        
        # Adatkinyerés
        tx_details = ntry.find('.//ns:NtryDtls/ns:TxDtls', ns)
        partner_name, remittance_info, add_info = "", "", ""
        
        add_info_node = ntry.find('ns:AddtlTxInf', ns)
        add_info = add_info_node.text if add_info_node is not None else ""

        if tx_details is not None:
            rltd_pties = tx_details.find('ns:RltdPties', ns)
            if rltd_pties is not None:
                cdtr = rltd_pties.find('.//ns:Cdtr/ns:Nm', ns)
                dbtr = rltd_pties.find('.//ns:Dbtr/ns:Nm', ns)
                partner_name = (cdtr.text if cdtr is not None else (dbtr.text if dbtr is not None else ""))
            
            rmt_node = tx_details.find('.//ns:RmtInf/ns:Ustrd', ns)
            remittance_info = rmt_node.text if rmt_node is not None else ""

        # Jogcímkód meghatározása
        kod = ""
        jogcim = ""
        megjegyzes = ""
        evszam = ""
        rem_low = remittance_info.lower()
        part_low = partner_name.lower()
        add_low = add_info.lower()

        if "output" in rem_low:
            transactions.append(add_new_line(sorsz, bookg_dt, bizonylat_sorszam, kod, jogcim, credit_amount, debit_amount, 
                                             current_running_balance, partner_name, remittance_info, "", evszam))
            sorsz += 1

            file_lista = os.listdir(folder_path)
            for i in file_lista:
                if rem_low[4:8] in i:
                    uj_adat = ujfFeldolgozó(f"{folder_path}/{i}")
                    if int(uj_adat[1]) >= int(credit_amount):
                        szamok = [x[1] for x in uj_adat[-1]]
                        for szam in range(0, len(szamok)+1):
                            eredmeny = osszeg_kereso(credit_amount, szam, szamok)
                            if eredmeny != None:
                                break

                        for bejegyzes in uj_adat[-1]:
                            if bejegyzes[-1] in eredmeny:
                                transactions.append(add_new_line(date=bookg_dt, biz=bizonylat_sorszam, kod=kod, jogc=jogcim, jovir=bejegyzes[1], kozlemeny=remittance_info, ev=evszam))
                    else:
                        for bejegyzes in uj_adat[-1]:
                            transactions.append(add_new_line(date=bookg_dt, biz=bizonylat_sorszam, kod=kod, jogc=jogcim, jovir=bejegyzes[1], 
                                                             egyenleg=current_running_balance, partner=partner_name, kozlemeny=remittance_info, ev=evszam))
    
        else:
            # 112: OTP Mobil / Persely
            if "otp mobil" in part_low or "simplepay" in part_low or "persely" in rem_low:
                kod = 112
            # 111: Egyházadó / Önkéntes hozzájárulás
            elif any(x in rem_low or x in add_low for x in ["adó", "egyházadó", "egyház adó", "önkéntes egyházi", "hozzájárulás", "egyhaz ado", "egyhazado", "onkentes", "egyhazi ado", "hozzáj", "egyházi adó"]):
                kod = 111
                evszam = datetime.now().year
                for i in rem_low.split():
                    try:
                        ev = int(i)
                        if ev > 2000 and ev <= datetime.now().year:
                            evszam = ev
                            rem_low -= i
                        break
                    except:
                        continue
                    finally:
                        name_entities = get_name_entities(remittance_info)
                        megjegyzes = megjegyzesKeszito(name_entities)
                        if len(megjegyzes) == 0:
                            name_entities = get_name_entities(partner_name.lower().title())
                            megjegyzes = megjegyzesKeszito(name_entities)

            elif "stóladíj" in rem_low:
                kod = 113
                megjegyzes = "Stóladíj"
            # 117: Orgona
            elif "orgona" in rem_low:
                kod = 117
                megjegyzes = "Adomány"
            # 151: Kamat
            elif any(x in add_low or x in rem_low for x in ["kifizetett kamat", "elszámolt kamat", "elszamolt kamat"]):
                kod = 151
                megjegyzes = "Kifizetett/Elszámolt Kamat"
            # 312: Telekommunikáció
            elif "magyar telekom" in part_low:
                kod = 312
            # 314: Ostya
            elif "ostyaellátó" in part_low or "ostyaellato" in part_low:
                kod = 314
            # 316: Közművek és biztonság
            elif any(x in part_low for x in ["elmű", "emász", "multi alarm", "szent györgy bizt", "nhkv", "mohu", "alarm electronic", "mvm next", "e.on"]):
                kod = 316
                megjegyzes = f"{partner_name} ({remittance_info})"
            # 319: Temetővel kapcsolatos közmű (Daköv/FlexiTon + temető szó)
            elif ("daköv" in part_low or "flexiton" in part_low) and "temet" in rem_low:
                kod = 319
                megjegyzes = "Temető"
            # 361: Banki jutalékok és díjak
            elif any(x in add_low or x in rem_low for x in ["jutalék", "pek ktg", "utalási díj", "tranzakciós illeték",
                                                            "állandó átutalás díja", "allando atutalas dija", "postai készpénzátutalás",
                                                            "pénzforgalmi", "jegy visszavásárlás", "bizományosi jutalék terhelése",
                                                            "bankközi átutalás", "bankközi atutalas"]):
                kod = 361
                if "ad" in rem_low and "giro" in rem_low:
                    megjegyzes = "utalási díj"
                else:
                    megjegyzes = "Banki költség"
            # 374: Oldallagos hozzájárulás
            elif "oldallagos" in rem_low and "plébániához" in rem_low:
                kod = 374
            # 411: Átfutó befizetés
            elif "átfutó" in rem_low or "átfutó" in add_low:
                kod = 411

            transactions.append(add_new_line(sorsz, bookg_dt, bizonylat_sorszam, kod, jogcim, credit_amount, debit_amount, 
                                             current_running_balance, partner_name, remittance_info, megjegyzes, evszam))
            sorsz += 1

    transactions.append(add_new_line(egyenleg=closing_balance, megj="Záró Egyenleg"))

    # DataFrame összeállítása
    df = pd.DataFrame(transactions)


    # Ha a kimeneti file nem létezik:
    if not os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=f"{datetime.now().strftime("%Y_%m")}")
            worksheet = writer.sheets[f"{datetime.now().strftime("%Y_%m")}"]

    with pd.ExcelWriter(output_file, engine='openpyxl', mode="a", if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name=f"{datetime.now().strftime("%Y_%m")}")
        worksheet = writer.sheets[f"{datetime.now().strftime("%Y_%m")}"]
        
        # Színek megadása
        blue_font = Font(color="0000FF")
        green_font = Font(color="008000")
        red_font = Font(color="FF0000") 

        # Pénz Formátum Megadása
        money_fmt = '#,##0 "Ft"'

        # Formátumok és Színek Alkalmazása
        for row in range(2, len(df) + 2):
            worksheet.cell(row=row, column=4).font = blue_font
            worksheet.cell(row=row, column=5).font = blue_font
            
            if worksheet.cell(row=row, column=6).value is not None:
                worksheet.cell(row=row, column=6).font = green_font
                worksheet.cell(row=row, column=6).number_format = money_fmt

            if worksheet.cell(row=row, column=7).value is not None:
                worksheet.cell(row=row, column=7).font = red_font
                worksheet.cell(row=row, column=7).number_format = money_fmt

            worksheet.cell(row=row, column=8).number_format = money_fmt

        # Cellák Vertikálisan Középre és Automatikus Szélesség Beállítása

        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if cell.value:
                        # Kiszámoljuk a cella tartalmának hosszát
                        curr_len = len(str(cell.value))
                        if curr_len > max_length:
                            max_length = curr_len
                except:
                    pass
            adjusted_width = max(max_length + 2, 10)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        worksheet.column_dimensions["D"].width = 5

        center_alignment_vertical = Alignment(vertical='center')
        for row in range(2, len(df) + 2):
            for d in range(1, 13):
                worksheet.cell(row=row, column=d).alignment = center_alignment_vertical

        # Adott Oszlopok Középre Igazítása
        center_alignment_horizontal = Alignment(horizontal='center')
        for row in range(2, len(df) + 2):
            for d in [1, 2, 3, 4, 6, 7, 8, 12, 13]:
                worksheet.cell(row=row, column=d).alignment = center_alignment_horizontal

        # Fejlécek Középreigazítása):
        max_r = worksheet.max_row
        for d in range(1, 13):
            worksheet.cell(row=1, column=d).alignment = center_alignment_horizontal
            worksheet.cell(row=2, column=d).alignment = center_alignment_horizontal
            worksheet.cell(row=max_r, column=d).alignment = center_alignment_horizontal

        # Félkövér Betűtípus Alkalmazása
        for cell in range(3, max_r):
            worksheet[f'B{cell}'].font = Font(bold=True)
            worksheet[f'H{cell}'].font = Font(bold=True)

        # Al Lekérdezések Dátumának Kicsinyítése
        for i in range(1, max_r):
            if worksheet[f"A{i}"].value == "":
                worksheet.cell(row=i, column=2).font = Font(size=9, bold=True)
                worksheet.cell(row=i, column=2).alignment = Alignment(horizontal='right')

        # Oszlopok Feltételes Elrejtése
        if printable:
            worksheet.column_dimensions['I'].hidden = True
            worksheet.column_dimensions['J'].hidden = True

        # Nyitó és Záró Egyenleg Sorok Formázása
        zold_kitoltes = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for c in range(1, 13):
            worksheet.cell(row=2, column=c).fill = zold_kitoltes
            worksheet.cell(row=max_r, column=c).fill = zold_kitoltes

    print(f"Fájl mentve ({"Nyomtatható" if printable == 1 else "Default"}): {output_file} ")

if __name__ == "__main__":
    # Parancssori argumentumok létrehozása, kezelése
    parser = argparse.ArgumentParser(description="Exceles Bankszámlakivonat Készítő")
    parser.add_argument("-n", "--nyomtat", type=int, help="0: Default; 1: Nyomtathato", default=0)
    parser.add_argument("-f", "--file", type=str, help="A .xml file elérési útvonala.", default="kivonat.xml")
    parser.add_argument("-m", "--mappa", type=str, help="Az UJF filokat tartalmazó mappa relatív útvonala.", default="./ujf")
    parser.add_argument("-c", "--cfile", type=str, help="A célfile elérési útvonala.", default="bankszamlakivonat.xlsx")
    args = parser.parse_args()
    
    print("Betöltés...")
    global ner_pipeline
    transformers_logging.set_verbosity_error()
    logging.disable_progress_bar()
    ner_pipeline = pipeline("ner", model="NYTK/named-entity-recognition-nerkor-hubert-hungarian")
    time.sleep(2)
    

    if args.file:
        convert_camt053_to_xlsx(f'{args.file}', output_file=args.cfile, printable=args.nyomtat, folder_path=args.mappa)
    else:
        convert_camt053_to_xlsx('kivonat.xml', 'bankszamlakivonat.xlsx', printable=args.nyomtat, folder_path=args.mappa)